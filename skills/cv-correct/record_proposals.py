"""Parse a subagent response and write its proposals + GPO issues to disk.

Tolerant JSON extraction: if the response file isn't pure JSON, look for the
first ``{`` ... matching ``}`` block. On total parse failure, log + exit 0
(the workflow should not halt on a single bad subagent response).

Writes:
- New entries appended to ``pending_corrections.json`` (deduped on
  ``(type, trigger, correction)``)
- New issue rows appended to ``gpo_issue_log.xlsx`` (header created if missing)
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

# Make the pipeline package importable.
PIPELINE_DIR = Path("/home/G39248410/citizen_voice/Code/data-preprocessing-pipeline")
sys.path.insert(0, str(PIPELINE_DIR))

from pipeline.corrections_registry import (  # noqa: E402
    CorrectionEntry,
    CorrectionsRegistry,
)

# GPO log columns — schema set by stakeholder workflow (the columns the user
# already tracks manually). Subagent issue keys map to these via _issue_to_row.
GPO_LOG_COLUMNS = [
    "Citable As",
    "Volume",
    "Congress",
    "issue",
    "Should_say",
    "text_says",
    "location_in_xml",
    "line_numbers",
]

# The ONLY keys an `issues[]` entry from the subagent may contain. Anything
# else is silently discarded at row-write time — the GPO log is frozen at
# these 8 columns and the subagent is not authorized to introduce new ones.
_ALLOWED_ISSUE_KEYS = {
    "citable_as", "volume", "congress", "issue",
    "should_say", "text_says", "location_in_xml", "line_numbers",
}


def _issue_to_row(issue: dict, default_vol: int) -> list:
    """Map a subagent issue object to the row in GPO_LOG_COLUMNS order.

    The subagent emits snake_case keys (per system_prompt.md). Excel cells
    have a 32,767-character hard cap — long xml fragments are truncated.

    Unknown keys are dropped silently and logged as a warning — the GPO log
    schema is frozen at exactly the 8 columns in ``GPO_LOG_COLUMNS``.
    """
    def _str(value) -> str:
        return "" if value is None else str(value)

    unknown = set(issue.keys()) - _ALLOWED_ISSUE_KEYS
    if unknown:
        logger.warning(
            "Dropping unknown keys from issue (only 8 GPO columns are allowed): %s",
            sorted(unknown),
        )

    text_says = _str(issue.get("text_says"))
    if len(text_says) > 32000:
        text_says = text_says[:32000] + "…<TRUNCATED>"

    return [
        _str(issue.get("citable_as")),                     # "Citable As"
        issue.get("volume", default_vol),                  # "Volume"
        _str(issue.get("congress")),                       # "Congress"
        _str(issue.get("issue")),                          # "issue"
        _str(issue.get("should_say")),                     # "Should_say"
        text_says,                                         # "text_says"
        _str(issue.get("location_in_xml")),                # "location_in_xml"
        _str(issue.get("line_numbers")),                   # "line_numbers"
    ]

logger = logging.getLogger("cv-correct.record_proposals")


def _extract_json_blob(text: str) -> dict | None:
    """Best-effort JSON extraction from a free-form response."""
    text = text.strip()
    if not text:
        return None
    # Common case: clean JSON
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # Strip ```json ... ``` fences
    fenced = re.search(r"```(?:json)?\s*(.+?)\s*```", text, re.DOTALL)
    if fenced:
        try:
            return json.loads(fenced.group(1))
        except json.JSONDecodeError:
            pass
    # Find the first balanced {} block
    start = text.find("{")
    if start < 0:
        return None
    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape:
            escape = False
            continue
        if ch == "\\":
            escape = True
            continue
        if ch == '"' and not escape:
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                candidate = text[start:i + 1]
                try:
                    return json.loads(candidate)
                except json.JSONDecodeError:
                    return None
    return None


def _proposals_to_entries(proposals: list[dict], vol: int, agent_version: str) -> list[CorrectionEntry]:
    now = datetime.now().isoformat()
    out: list[CorrectionEntry] = []
    for p in proposals:
        if not isinstance(p, dict):
            continue
        ptype = p.get("type", "other")
        if ptype not in ("law_id", "section_number", "other"):
            logger.warning("Skipping proposal with unknown type=%r", ptype)
            continue
        entry = CorrectionEntry.from_dict({
            "id": 0,
            "type": ptype,
            "trigger": p.get("trigger", {}) or {},
            "correction": p.get("correction", {}) or {},
            "evidence": p.get("evidence", {}) or {},
            "proposed_at": now,
            "discovered_in_vol": vol,
            "agent_version": agent_version,
            "confidence": p.get("confidence"),
            "applied_in_runs": [{"vol": vol, "run_ts": now}],
            "seen_again_count": 0,
            "status": "pending",
            "reviewer": None,
            "review_note": None,
            "reviewed_at": None,
        })
        out.append(entry)
    return out


def _issue_dedup_key(row: list) -> tuple:
    """Uniqueness key for a GPO row.

    A row is uniquely identified by (Volume, Citable As, line_numbers,
    issue). This means:

    - **Same anomaly re-discovered on a re-run** → same key → skipped, even
      if other columns drift slightly.
    - **Same correction rule affecting two distinct PLs** → different
      ``Citable As`` (each PL's own Statutes citation) → both rows kept.
    - **Two distinct issues at the same line range** → different ``issue``
      text → both rows kept.

    Columns order: [Citable As, Volume, Congress, issue, Should_say,
    text_says, location_in_xml, line_numbers].
    """
    citable_as, volume, _congress, issue, _should, _says, _loc, line_numbers = row
    return (
        str(volume).strip(),
        str(citable_as).strip().casefold(),
        str(line_numbers).strip(),
        str(issue).strip().casefold(),
    )


def _append_to_gpo_log(issues: list[dict], gpo_log_path: Path, vol: int) -> tuple[int, int]:
    """Append issue rows, de-duplicating against rows already in the log.

    Returns ``(written, skipped_as_dup)``.

    Dedup is per-volume — each unique (volume, citable_as, line_numbers,
    issue) tuple appears at most once. A re-run on the same volume that
    surfaces the same anomalies is therefore a no-op for the GPO log.
    """
    if not issues:
        return 0, 0
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as e:
        logger.error("openpyxl missing — cannot write GPO log: %s", e)
        return 0, 0

    if gpo_log_path.exists():
        wb = load_workbook(gpo_log_path)
        ws = wb.active
        existing_keys = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if len(row) < len(GPO_LOG_COLUMNS):
                continue
            existing_keys.add(_issue_dedup_key(list(row[: len(GPO_LOG_COLUMNS)])))
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(GPO_LOG_COLUMNS)
        existing_keys = set()

    written = 0
    skipped = 0
    for issue in issues:
        if not isinstance(issue, dict):
            continue
        row = _issue_to_row(issue, default_vol=vol)
        key = _issue_dedup_key(row)
        if key in existing_keys:
            skipped += 1
            continue
        ws.append(row)
        existing_keys.add(key)
        written += 1

    gpo_log_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(gpo_log_path)
    return written, skipped


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--volume", type=int, required=True)
    parser.add_argument("--response", required=True, type=Path,
                        help="Path to the file containing the subagent's text response.")
    parser.add_argument("--pending", required=True, type=Path)
    parser.add_argument("--gpo-log", required=True, type=Path)
    parser.add_argument("--candidates", type=Path, default=None,
                        help="candidates_N.json — used to verify every anomaly "
                             "the prefilter surfaced is represented in the GPO log.")
    parser.add_argument("--agent-version", default="cv-correct@0.1.0")
    parser.add_argument("--log-level", default="INFO",
                        choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    args = parser.parse_args(argv)

    logging.basicConfig(level=args.log_level, format="%(levelname)s %(name)s: %(message)s")

    if not args.response.exists():
        logger.warning("Response file not found: %s — nothing to record", args.response)
        return 0

    raw = args.response.read_text()
    parsed = _extract_json_blob(raw)
    if parsed is None:
        logger.warning("Could not extract JSON from %s — recording nothing", args.response)
        return 0

    proposals = parsed.get("proposals") or []
    issues = parsed.get("issues") or []
    notes = parsed.get("notes") or ""

    # Mandatory-coverage check — the subagent must emit one issue per anomaly
    # the prefilter surfaced. Under-emission is a quality flag, not a hard
    # error: we still record what was returned, but the operator sees a warning.
    if args.candidates and args.candidates.exists():
        try:
            cand = json.loads(args.candidates.read_text())
            n_anomalies = (
                len(cand.get("law_number_anomalies") or [])
                + len(cand.get("unique_key_duplicates") or [])
            )
            if len(issues) < n_anomalies:
                logger.warning(
                    "vol%d: GPO log under-emission — %d anomalies surfaced by "
                    "prefilter, only %d issues[] returned by subagent. Every "
                    "anomaly must produce a GPO row (system_prompt.md invariant).",
                    args.volume, n_anomalies, len(issues),
                )
        except (OSError, json.JSONDecodeError) as e:
            logger.warning("Could not load candidates for coverage check: %s", e)

    # Pending registry append (deduped by (type, trigger, correction)).
    output_dir = args.pending.parent
    registry = CorrectionsRegistry(output_dir)
    entries = _proposals_to_entries(proposals, args.volume, args.agent_version)
    novel = registry.append_pending(entries)

    # GPO log append (per-volume dedup on (volume, citable_as, line_numbers, issue)).
    written, skipped = _append_to_gpo_log(issues, args.gpo_log, args.volume)

    print(
        f"vol{args.volume}: {len(proposals)} proposals in response, {novel} novel; "
        f"{len(issues)} issues, {written} appended to GPO log, {skipped} "
        f"skipped as duplicates of existing rows. notes={notes!r}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
